from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import sys
from pathlib import Path
from typing import Any
from urllib.parse import quote

import chardet
from docx import Document as DocxDocument
from dotenv import load_dotenv
from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook
from pypdf import PdfReader

from crisai import ms_graph
from crisai.config import load_settings
from crisai.logging_utils import append_json_log_line, configure_mcp_framework_logging
from crisai.powerpoint import extract_powerpoint_from_bytes

load_dotenv()

mcp = FastMCP("crisai-sharepoint")

ROOT = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else Path.cwd().resolve()
ROOT.mkdir(parents=True, exist_ok=True)

LOG_FILE = load_settings().log_dir / "sharepoint_mcp.log"
CACHE_DIR = ROOT / ".auth"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

SUPPORTED_TEXT_SUFFIXES = {".txt", ".md", ".json", ".yaml", ".yml", ".py", ".log", ".csv"}
SUPPORTED_DOC_SUFFIXES = {".docx", ".pdf", ".pptx", ".xlsx"}
_READ_HANDLE_PREFIX = "sharepoint_doc:"
_REF_PREFIX = "spdoc-"

# Maps a short reference (spdoc-XXXXXXXX) to the real (drive_id, item_id). A
# Graph drive id is ~66 chars of opaque base64 that the model regularly corrupts
# when it transcribes it across the search->read handover (the recurring "404 on
# a malformed drive id" failure). Every search/list result instead carries a
# short, stable, copy-safe ref the server resolves here, so the model never has
# to copy a long id.
_REF_REGISTRY: dict[str, tuple[str, str]] = {}


def _mint_ref(drive_id: str, item_id: str) -> str:
    """Return a short, stable read reference for a drive/item pair and register it."""
    # Non-cryptographic: a short collision-resistant id for a copy-safe ref token.
    digest = hashlib.sha1(f"{drive_id}:{item_id}".encode(), usedforsecurity=False).hexdigest()[:8]
    ref = f"{_REF_PREFIX}{digest}"
    _REF_REGISTRY[ref] = (drive_id, item_id)
    return ref


def _encode_read_handle(drive_id: str, item_id: str) -> str:
    """Encode a SharePoint drive/item pair as a short, copy-safe read handle."""
    return _mint_ref(drive_id, item_id)


def _decode_read_handle(read_handle: str) -> tuple[str, str]:
    """Resolve a SharePoint read handle to its (drive_id, item_id).

    The primary form is a short server-minted ref; a legacy base64 handle is
    still accepted so older references keep working.
    """
    handle = (read_handle or "").strip()
    if handle in _REF_REGISTRY:
        return _REF_REGISTRY[handle]
    if handle.startswith(_REF_PREFIX):
        raise ValueError(
            f"Unknown SharePoint reference {handle!r}. Re-run the search or list to get current references."
        )
    if handle.startswith(_READ_HANDLE_PREFIX):
        token = handle[len(_READ_HANDLE_PREFIX) :]
        try:
            padded = token + ("=" * (-len(token) % 4))
            payload = json.loads(base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8"))
        except (ValueError, json.JSONDecodeError) as exc:
            raise ValueError("Malformed SharePoint read_handle.") from exc
        drive_id = str(payload.get("drive_id") or "")
        item_id = str(payload.get("item_id") or "")
        if not drive_id or not item_id:
            raise ValueError("SharePoint read_handle is missing drive_id or item_id.")
        return drive_id, item_id
    raise ValueError("read_handle must be a SharePoint reference from a search or list result.")


def _configure_mcp_logging() -> None:
    """Keep MCP framework INFO logs out of the interactive CLI.

    Warnings and errors are still written to this server log file.
    """
    configure_mcp_framework_logging(LOG_FILE, service_component="sharepoint_mcp")


def log_event(message: str) -> None:
    append_json_log_line(
        LOG_FILE,
        message,
        logger_name="crisai.mcp.sharepoint",
        service_component="sharepoint_mcp",
    )


ms_graph.configure_workspace(ROOT, namespace="sharepoint")
ms_graph.set_telemetry_hook(log_event)


# Read tools must never trigger an interactive login. Without silent_only the
# Graph helper falls back to a device-code flow that blocks until the user
# completes it manually, so an unauthenticated call hangs until the MCP client
# timeout fires (240s) and crashes the whole agent run. silent_only=True instead
# raises immediately with a "call login_sharepoint" hint the agent can act on.
# Interactive login stays confined to the explicit login_sharepoint tool.
def _graph_get(path: str, params: dict[str, Any] | None = None, timeout: int = 60) -> dict[str, Any]:
    return ms_graph.graph_get(path, params=params, timeout=timeout, silent_only=True)


def _graph_get_bytes(url: str) -> bytes:
    return ms_graph.graph_get_bytes(url, silent_only=True)


def _detect_text_encoding(data: bytes) -> str:
    detected = chardet.detect(data)
    return detected.get("encoding") or "utf-8"


def _read_text_like_bytes(data: bytes) -> str:
    encoding = _detect_text_encoding(data)
    try:
        return data.decode(encoding)
    except Exception:
        return data.decode("utf-8", errors="replace")


def _read_csv_bytes(data: bytes, max_rows: int = 100) -> str:
    encoding = _detect_text_encoding(data)
    text = data.decode(encoding, errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows: list[list[str]] = []
    for idx, row in enumerate(reader):
        if idx >= max_rows:
            rows.append(["... truncated ..."])
            break
        rows.append([str(cell) for cell in row])
    return "\n".join(" | ".join(row) for row in rows)


def _read_docx_bytes(data: bytes) -> str:
    with io.BytesIO(data) as bio:
        doc = DocxDocument(bio)
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    return "\n\n".join(paragraphs)


def _read_pdf_bytes(data: bytes) -> str:
    with io.BytesIO(data) as bio:
        reader = PdfReader(bio)
        pages = []
        for i, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if text:
                pages.append(f"[Page {i}]\n{text}")
    return "\n\n".join(pages)


def _read_pptx_bytes(data: bytes) -> str:
    return extract_powerpoint_from_bytes(data).to_text()


def _read_xlsx_bytes(data: bytes, max_rows: int = 50, max_cols: int = 20) -> str:
    with io.BytesIO(data) as bio:
        wb = load_workbook(filename=bio, read_only=True, data_only=True)
        out = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            out.append(f"[Sheet: {sheet_name}]")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if row_count >= max_rows:
                    out.append("... truncated ...")
                    break
                values = []
                for cell in row[:max_cols]:
                    values.append("" if cell is None else str(cell))
                if any(v != "" for v in values):
                    out.append(" | ".join(values))
                    row_count += 1
            out.append("")
    return "\n".join(out).strip()


def _extract_bytes_by_suffix(data: bytes, suffix: str) -> str:
    suffix = suffix.lower()

    if suffix in {".txt", ".md", ".json", ".yaml", ".yml", ".py", ".log"}:
        return _read_text_like_bytes(data)

    if suffix == ".csv":
        return _read_csv_bytes(data)

    if suffix == ".docx":
        return _read_docx_bytes(data)

    if suffix == ".pdf":
        return _read_pdf_bytes(data)

    if suffix == ".pptx":
        return _read_pptx_bytes(data)

    if suffix == ".xlsx":
        return _read_xlsx_bytes(data)

    raise ValueError(f"Unsupported document type: {suffix or 'no extension'}")


def _is_likely_personal_onedrive_site(site: dict[str, Any]) -> bool:
    """Return True when a site row looks like a user's personal OneDrive library.

    Team SharePoint sites usually live under ``{tenant}.sharepoint.com/sites/...``.
    OneDrive for Business personal sites commonly use ``*-my.sharepoint.com/personal/...``.
    """
    web = str(site.get("webUrl") or "").lower()
    return "-my.sharepoint.com" in web and "/personal/" in web


def _normalise_item(item: dict[str, Any]) -> dict[str, Any]:
    parent = item.get("parentReference", {}) or {}
    file_info = item.get("file", {}) or {}
    folder_info = item.get("folder", {}) or {}

    web_url = item.get("webUrl")
    item_id = str(item.get("id") or "")
    drive_id = str(parent.get("driveId") or "")
    read_handle = _encode_read_handle(drive_id, item_id) if drive_id and item_id else ""
    return {
        "id": item.get("id"),
        "name": item.get("name"),
        "webUrl": web_url,
        # Explicit alias so agents surface a browser-openable link for OneDrive/SharePoint.
        "open_url": web_url,
        "size": item.get("size"),
        "createdDateTime": item.get("createdDateTime"),
        "lastModifiedDateTime": item.get("lastModifiedDateTime"),
        "mimeType": file_info.get("mimeType"),
        "isFolder": bool(folder_info),
        "parentPath": parent.get("path"),
        "driveId": parent.get("driveId"),
        "read_handle": read_handle,
    }

@mcp.tool()
def login_sharepoint() -> str:
    """Force interactive Microsoft login and cache tokens for later SharePoint calls."""
    log_event("login_sharepoint")
    ms_graph.acquire_token(force_interactive=True)
    info = ms_graph.read_token_info()
    return f"SharePoint login successful. Account={info.get('account')} Scopes={info.get('scope')}"


@mcp.tool()
def sharepoint_auth_status() -> dict[str, Any]:
    """Return SharePoint auth status without forcing interactive login."""
    log_event("sharepoint_auth_status")
    return ms_graph.delegated_auth_status()

@mcp.tool()
def who_am_i() -> dict[str, Any]:
    """Return basic information about the signed-in Microsoft Graph user."""
    log_event("who_am_i")
    return _graph_get("/me")


@mcp.tool()
def list_sites(query: str = "*", max_hits: int = 10) -> list[dict[str, Any]]:
    log_event(f"list_sites query={query!r} max_hits={max_hits}")
    data = _graph_get("/sites", params={"search": query}, timeout=90)
    values = data.get("value", [])[:max_hits]
    return [
        {
            "id": site.get("id"),
            "name": site.get("name"),
            "displayName": site.get("displayName"),
            "webUrl": site.get("webUrl"),
            "open_url": site.get("webUrl"),
            "description": site.get("description"),
            "hostname": (site.get("siteCollection") or {}).get("hostname"),
        }
        for site in values
    ]


@mcp.tool()
def list_my_drives() -> list[dict[str, Any]]:
    """List drives visible under the signed-in user's account (includes personal OneDrive).

    For **SharePoint team/site** file search, prefer ``search_sharepoint_site_documents``
    or ``list_sites`` + ``search_site_drive_documents`` instead of searching only here,
    otherwise results skew toward OneDrive.
    """
    log_event("list_my_drives")
    data = _graph_get("/me/drives")
    return [
        {
            "id": drive.get("id"),
            "name": drive.get("name"),
            "description": drive.get("description"),
            "webUrl": drive.get("webUrl"),
            "open_url": drive.get("webUrl"),
            "driveType": drive.get("driveType"),
        }
        for drive in data.get("value", [])
    ]


@mcp.tool()
def list_site_drives(site_id: str) -> list[dict[str, Any]]:
    """List drives in a SharePoint site."""
    log_event(f"list_site_drives site_id={site_id}")
    data = _graph_get(f"/sites/{site_id}/drives")
    return [
        {
            "id": drive.get("id"),
            "name": drive.get("name"),
            "description": drive.get("description"),
            "webUrl": drive.get("webUrl"),
            "open_url": drive.get("webUrl"),
            "driveType": drive.get("driveType"),
        }
        for drive in data.get("value", [])
    ]


@mcp.tool()
def list_drive_items(drive_id: str, item_id: str = "root", max_items: int = 50) -> list[dict[str, Any]]:
    """List items under a drive folder. Use item_id='root' for the drive root."""
    log_event(f"list_drive_items drive_id={drive_id} item_id={item_id} max_items={max_items}")
    data = _graph_get(f"/drives/{drive_id}/items/{item_id}/children")
    values = data.get("value", [])[:max_items]
    return [_normalise_item(item) for item in values]


# Drive-id values that mean "the user's own OneDrive" rather than a real Graph
# drive id. Agents frequently emit these instead of threading the id from
# list_my_drives, so resolve them to the primary drive rather than failing.
_PERSONAL_DRIVE_ALIASES = frozenset(
    {"", "root", "me", "my", "mine", "onedrive", "default", "current", "__placeholder__", "__invalid__"}
)


def _my_drive_id() -> str:
    """Return the signed-in user's primary OneDrive drive id."""
    data = _graph_get("/me/drive")
    drive_id = data.get("id")
    if not drive_id:
        raise RuntimeError("Could not resolve the signed-in user's OneDrive drive.")
    return str(drive_id)


def _resolve_drive_id(drive_id: str) -> str:
    """Map an empty or placeholder drive id to the user's primary OneDrive."""
    candidate = (drive_id or "").strip()
    looks_placeholder = (
        not candidate
        or candidate.lower() in _PERSONAL_DRIVE_ALIASES
        # Any __SOMETHING__ token the model invents (e.g. __TODO__, __drive_id__).
        or (candidate.startswith("__") and candidate.endswith("__"))
    )
    return _my_drive_id() if looks_placeholder else candidate


@mcp.tool()
def search_my_onedrive(query: str, max_hits: int = 20) -> list[dict[str, Any]]:
    """Search the signed-in user's personal OneDrive by filename or content.

    Use this for "in my OneDrive" requests: it targets the user's primary drive
    directly, so you do **not** need a drive_id from ``list_my_drives``.
    """
    log_event(f"search_my_onedrive query={query!r} max_hits={max_hits}")
    encoded = quote(query, safe="")
    data = _graph_get(f"/me/drive/root/search(q='{encoded}')", timeout=90)
    values = data.get("value", [])[:max_hits]
    return [_normalise_item(item) for item in values]


@mcp.tool()
def search_drive_documents(drive_id: str, query: str, max_hits: int = 20) -> list[dict[str, Any]]:
    """Search files in a drive. If drive_id is empty or a placeholder (for
    example 'root' or 'me'), the user's personal OneDrive is searched."""
    log_event(f"search_drive_documents drive_id={drive_id} query={query!r} max_hits={max_hits}")
    resolved = _resolve_drive_id(drive_id)
    encoded = quote(query, safe="")
    data = _graph_get(f"/drives/{resolved}/root/search(q='{encoded}')", timeout=90)
    values = data.get("value", [])[:max_hits]
    return [_normalise_item(item) for item in values]

@mcp.tool()
def search_site_drive_documents(site_id: str, query: str, max_hits: int = 20) -> list[dict[str, Any]]:
    log_event(f"search_site_drive_documents site_id={site_id} query={query!r} max_hits={max_hits}")
    encoded = quote(query, safe="")
    data = _graph_get(f"/sites/{site_id}/drive/root/search(q='{encoded}')", timeout=90)
    values = data.get("value", [])[:max_hits]
    return [_normalise_item(item) for item in values]


@mcp.tool()
def search_sharepoint_site_documents(
    document_query: str,
    site_search: str = "*",
    max_sites: int = 12,
    max_hits_per_site: int = 10,
    include_onedrive_personal_sites: bool = False,
) -> list[dict[str, Any]]:
    """Search SharePoint **team/site** document libraries, not personal OneDrive by default.

    Use this when the user asks for **SharePoint** (sites/libraries) and does **not**
    restrict the request to personal **OneDrive** only. It discovers sites via
    ``list_sites`` and runs ``search_site_drive_documents`` on each site's default
    library.

    Do **not** satisfy SharePoint-only requests using only ``list_my_drives`` plus
    ``search_drive_documents``, because that path is usually the user's OneDrive.

    Args:
        document_query: Text for Graph drive item search (for example a topic name).
        site_search: Query passed to ``list_sites`` (use ``*`` or a site keyword).
        max_sites: Maximum number of sites to search.
        max_hits_per_site: Maximum hits returned per site library.
        include_onedrive_personal_sites: When True, include ``*-my.sharepoint.com/personal/`` sites.
    """
    log_event(
        "search_sharepoint_site_documents "
        f"q={document_query!r} site_search={site_search!r} max_sites={max_sites} "
        f"max_hits_per_site={max_hits_per_site} include_onedrive={include_onedrive_personal_sites}"
    )
    sites = list_sites(query=site_search, max_hits=max_sites)
    aggregated: list[dict[str, Any]] = []
    max_total = 80

    for site in sites:
        if len(aggregated) >= max_total:
            break
        if not include_onedrive_personal_sites and _is_likely_personal_onedrive_site(site):
            continue
        site_id = site.get("id")
        if not site_id:
            continue
        hits = search_site_drive_documents(
            site_id=str(site_id),
            query=document_query,
            max_hits=max_hits_per_site,
        )
        for item in hits:
            if len(aggregated) >= max_total:
                break
            row = dict(item)
            row["site_display_name"] = site.get("displayName")
            row["site_web_url"] = site.get("webUrl")
            row["site_id"] = site_id
            aggregated.append(row)

    return aggregated


# Internal: not exposed as a tool. Reads go through the by-handle tools, which
# resolve a short ref to the real ids, so the model never passes a raw drive id.
def get_sharepoint_document_metadata(drive_id: str, item_id: str) -> dict[str, Any]:
    """Return metadata for a SharePoint drive item."""
    log_event(f"get_sharepoint_document_metadata drive_id={drive_id} item_id={item_id}")
    item = _graph_get(f"/drives/{drive_id}/items/{item_id}")
    return _normalise_item(item)


@mcp.tool()
def get_sharepoint_document_metadata_by_handle(read_handle: str) -> dict[str, Any]:
    """Return metadata for a SharePoint drive item using an opaque read handle."""
    log_event("get_sharepoint_document_metadata_by_handle")
    drive_id, item_id = _decode_read_handle(read_handle)
    return get_sharepoint_document_metadata(drive_id=drive_id, item_id=item_id)


# Internal helper (see get_sharepoint_document_metadata): reached via the
# by-handle read tool only, never with a model-supplied raw drive id.
def read_sharepoint_document(drive_id: str, item_id: str) -> str:
    """Download and extract text from a supported SharePoint document."""
    log_event(f"read_sharepoint_document drive_id={drive_id} item_id={item_id}")

    item = _graph_get(f"/drives/{drive_id}/items/{item_id}")
    name = str(item.get("name") or "")
    suffix = Path(name).suffix.lower()

    if not suffix:
        raise ValueError(f"Cannot determine file type for item name: {name}")

    content_url = f"{ms_graph.GRAPH_BASE}/drives/{drive_id}/items/{item_id}/content"
    data = _graph_get_bytes(content_url)

    extracted = _extract_bytes_by_suffix(data, suffix)
    return extracted if extracted.strip() else f"[No readable text extracted from {name}]"


# Internal helper (see get_sharepoint_document_metadata): reached via the
# by-handle inspect tool only, never with a model-supplied raw drive id.
def inspect_sharepoint_powerpoint(drive_id: str, item_id: str) -> dict[str, Any]:
    """Download a SharePoint PowerPoint and return structured slide extraction data."""
    log_event(f"inspect_sharepoint_powerpoint drive_id={drive_id} item_id={item_id}")

    item = _graph_get(f"/drives/{drive_id}/items/{item_id}")
    name = str(item.get("name") or "")
    suffix = Path(name).suffix.lower()
    if suffix != ".pptx":
        raise ValueError(f"Expected a .pptx file, got: {suffix or 'no extension'}")

    content_url = f"{ms_graph.GRAPH_BASE}/drives/{drive_id}/items/{item_id}/content"
    extraction = extract_powerpoint_from_bytes(_graph_get_bytes(content_url))
    result = extraction.to_dict()
    result["source"] = _normalise_item(item)
    return result


@mcp.tool()
def inspect_sharepoint_powerpoint_by_handle(read_handle: str) -> dict[str, Any]:
    """Inspect a SharePoint PowerPoint using an opaque read handle."""
    log_event("inspect_sharepoint_powerpoint_by_handle")
    drive_id, item_id = _decode_read_handle(read_handle)
    return inspect_sharepoint_powerpoint(drive_id=drive_id, item_id=item_id)


@mcp.tool()
def read_sharepoint_document_by_handle(read_handle: str) -> str:
    """Download and extract text from a supported SharePoint document read handle."""
    log_event("read_sharepoint_document_by_handle")
    drive_id, item_id = _decode_read_handle(read_handle)
    return read_sharepoint_document(drive_id=drive_id, item_id=item_id)


_configure_mcp_logging()

if __name__ == "__main__":
    mcp.run()
